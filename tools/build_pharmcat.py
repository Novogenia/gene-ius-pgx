# -*- coding: utf-8 -*-
"""
Baut den JS-Datenblock fuer das echte Genprofil aus den PharmCAT-Ausgaben und
die Leitlinien-Matrix aus 'Pharmgkb drug recommendations V4.xlsx'.

Quellen:
  pharmcat.phenotype (1).json   Diplotyp, Phaenotyp, Aktivitaetsscore je Gen
  pharmcat.match (1).json       gerufene und fehlende Positionen je Gen
  Pharmgkb drug recommendations V4.xlsx  Empfehlungstexte je Genotyp-Kombination

Regel: nichts erfinden. Was PharmCAT nicht ruft, bleibt 'nicht bestimmbar'.
Insbesondere wird kein CYP2D6-Diplotyp geraten, obwohl das 46 der 103
Empfehlungszeilen betrifft.

Ausgabe: pharmcat_profil.js (rein ASCII)
"""
import json, os, re, sys
from collections import defaultdict, Counter
from openpyxl import load_workbook

DL = r"C:\Users\DanielWallerstorfer\Downloads"
XL = r"C:\Users\DanielWallerstorfer\Novogenia GmbH\AI RESOURCES - Dokumente\PHARMACOGENETICS\Pharmgkb drug recommendations V4.xlsx"
OUT = "pharmcat_profil.js"

# ---------------------------------------------------------------- Hilfsmittel
def esc(s):
    """ASCII-sicher: Umlaute und Sonderzeichen als HTML-Entities."""
    if s is None: return ""
    s = re.sub(r"\s+", " ", str(s)).strip()
    rep = {"\u00e4":"&auml;","\u00f6":"&ouml;","\u00fc":"&uuml;","\u00c4":"&Auml;",
           "\u00d6":"&Ouml;","\u00dc":"&Uuml;","\u00df":"&szlig;",
           "\u2013":"&mdash;","\u2014":"&mdash;","\u2019":"&rsquo;","\u2018":"&lsquo;",
           "\u201c":"&ldquo;","\u201d":"&ldquo;","\u201e":"&bdquo;","\u2265":"&ge;",
           "\u2264":"&le;","\u2192":"&rarr;","\u00d7":"x","\u00b7":"&middot;",
           "\u00e9":"&eacute;","\u00e8":"&egrave;","\u00b0":"&deg;","\u00a0":" ",
           "\u2011":"-","\u00b5":"&micro;","\u2018":"'"}
    for a, b in rep.items(): s = s.replace(a, b)
    s = s.replace("\\", "").replace("`", "'").replace("${", "$ {")
    return "".join(c if ord(c) < 128 else "" for c in s)

def j(o): return json.dumps(o, separators=(",", ":"), ensure_ascii=True)

# Phaenotyp (PharmCAT/CPIC englisch) -> Kanon-Code
CANON = {
 "ultrarapid metabolizer":"UM","rapid metabolizer":"RM","normal metabolizer":"NM",
 "likely intermediate metabolizer":"IM","intermediate metabolizer":"IM",
 "likely poor metabolizer":"PM","poor metabolizer":"PM",
 "increased function":"IF","normal function":"NF","possible decreased function":"DF",
 "decreased function":"DF","poor function":"PF",
 "normal":"NM","deficient":"PM","variable":"?","deficient with cnshd":"PM",
 "indeterminate":"?","no result":"?","n/a":"?","":"?",
}
# Spreadsheet-Vokabular -> derselbe Kanon
SHEET = {"ULTRARAPID":"UM","RAPID":"RM","EXTENSIVE":"NM","NORMAL":"NM",
         "INTERMEDIATE":"IM","POOR":"PM"}
# Transporter-Funktion und Metabolisierer sind im Spreadsheet gleich benannt
AEQ = {"NF":"NM","DF":"IM","PF":"PM","IF":"UM"}
# Kanon -> Stufe der App (0 langsam ... 3 ultraschnell, -1 nicht bestimmbar)
LVLOF = {"PM":0,"PF":0,"IM":1,"DF":1,"NM":2,"NF":2,"RM":3,"UM":3,"IF":3}

# Deutsche Bezeichnung je Kanon-Code, getrennt fuer Enzyme und Transporter
DE_ENZ = {"UM":"Ultraschneller Metabolisierer","RM":"Schneller Metabolisierer",
          "NM":"Normaler Metabolisierer","IM":"Intermedi\u00e4rer Metabolisierer",
          "PM":"Langsamer Metabolisierer"}
DE_TRANS = {"IF":"Gesteigerte Transportfunktion","NF":"Normale Transportfunktion",
            "DF":"Verminderte Transportfunktion","PF":"Stark verminderte Transportfunktion"}
# G6PD ist kein Metabolisierer, sondern ein Schutzenzym der roten Blutkoerperchen.
# CPIC fuehrt dort Normal / Deficient / Deficient with CNSHA / Variable.
DE_SPEZIAL = {
 "G6PD": {"NM":"Kein G6PD-Mangel", "PM":"G6PD-Mangel", "IM":"G6PD-Mangel (teilweise)"},
}

# Art des Gens - bestimmt Wortwahl und ob eine Metabolisierer-Skala sinnvoll ist
KIND = {
 "CYP2B6":"enz","CYP2C19":"enz","CYP2C9":"enz","CYP2D6":"enz","CYP3A4":"enz",
 "CYP3A5":"enz","DPYD":"enz","NAT2":"enz","NUDT15":"enz","TPMT":"enz","UGT1A1":"enz",
 "SLCO1B1":"trans","ABCG2":"trans","ABCB1":"trans",
 "VKORC1":"ziel","CYP4F2":"enz","G6PD":"enz",
 "RYR1":"risiko","CACNA1S":"risiko","CFTR":"ziel",
 "HLA-A":"hla","HLA-B":"hla","IFNL3":"marker","MT-RNR1":"risiko",
}
# Wofuer das Gen klinisch steht - eine Zeile, aus CPIC-Leitlinienthemen
ROLLE = {
 "CYP2B6":"Abbau von Efavirenz und weiteren HIV-Wirkstoffen",
 "CYP2C19":"Abbau von Clopidogrel, Protonenpumpenhemmern und Antidepressiva",
 "CYP2C9":"Abbau von Gerinnungshemmern, Phenytoin und Schmerzmitteln",
 "CYP2D6":"Abbau von etwa jedem vierten Wirkstoff \u2013 Opioide, Antidepressiva, Betablocker",
 "CYP3A4":"Abbau sehr vieler Wirkstoffe, gemeinsam mit CYP3A5",
 "CYP3A5":"Abbau von Tacrolimus nach Organtransplantation",
 "CYP4F2":"Vitamin-K-Umsatz, beeinflusst die Warfarin-Dosis",
 "DPYD":"Abbau der Krebsmedikamente Fluorouracil und Capecitabin",
 "G6PD":"Schutz der roten Blutk\u00f6rperchen vor oxidativem Stress",
 "NAT2":"Abbau von Isoniazid und weiteren Wirkstoffen",
 "NUDT15":"Abbau von Thiopurinen wie Azathioprin",
 "TPMT":"Abbau von Thiopurinen wie Azathioprin und Mercaptopurin",
 "UGT1A1":"Ausscheidung von Irinotecan und Atazanavir",
 "SLCO1B1":"Aufnahme von Statinen in die Leber",
 "ABCG2":"Transport von Rosuvastatin und Allopurinol",
 "VKORC1":"Angriffspunkt der Vitamin-K-Antagonisten Warfarin und Acenocoumarol",
 "CFTR":"Chloridkanal \u2013 Ansprechen auf Ivacaftor bei Mukoviszidose",
 "RYR1":"Risiko einer malignen Hyperthermie unter Narkose",
 "CACNA1S":"Risiko einer malignen Hyperthermie unter Narkose",
 "HLA-A":"Risiko schwerer Hautreaktionen auf Carbamazepin",
 "HLA-B":"Risiko schwerer \u00dcberempfindlichkeit auf Abacavir und Carbamazepin",
 "IFNL3":"Ansprechen auf Interferon-Therapie bei Hepatitis C",
 "MT-RNR1":"Risiko einer Schwerh\u00f6rigkeit unter Aminoglykosid-Antibiotika",
}
# Allelfunktion englisch -> deutsch (Wortlaut wie in der App)
FN_DE = {
 "normal function":"normal","increased function":"gesteigert",
 "decreased function":"reduziert","no function":"keine",
 "possible decreased function":"m\u00f6glicherweise reduziert",
 "uncertain function":"unklar","unknown function":"unbekannt",
 "iv/normal":"normal","normal":"normal","deficient":"Mangel",
 "reference":"normal","null":"unbekannt","none":"unbekannt",
}

def fn_de(s):
    if s is None: return "unbekannt"
    return FN_DE.get(str(s).strip().lower(), esc(s))

def canon(p):
    return CANON.get((p or "").strip().lower(), "?")

# ---------------------------------------------------------------- PharmCAT
ph = json.load(open(os.path.join(DL, "pharmcat.phenotype (1).json"), encoding="utf-8"))
ma = json.load(open(os.path.join(DL, "pharmcat.match (1).json"), encoding="utf-8"))
mrows = {r["gene"]: r for r in ma["results"]}

genes = []
for g in sorted(ph["geneReports"]):
    r = ph["geneReports"][g]
    dl = r.get("recommendationDiplotypes") or []
    codes, phen, labels = [], [], []
    a1 = a2 = f1 = f2 = ""
    score = ""
    for d in dl:
        for p in (d.get("phenotypes") or []):
            c = canon(p)
            if c not in codes: codes.append(c)
            if p not in phen: phen.append(p)
        if d.get("label"): labels.append(d["label"])
        if d.get("activityScore") not in (None, "", "n/a") and not score:
            score = str(d["activityScore"])
    if len(dl) == 1:
        d = dl[0]
        al1, al2 = d.get("allele1") or {}, d.get("allele2") or {}
        a1, f1 = al1.get("name") or "", fn_de(al1.get("function"))
        a2, f2 = al2.get("name") or "", fn_de(al2.get("function"))
        # G6PD und MT-RNR1 sind haploid/X-chromosomal notiert - nur eine Kopie
        if not a2: a2, f2 = "", ""
    mr = mrows.get(g) or {}
    md = mr.get("matchData") or {}
    da = len(mr.get("variants") or [])
    fehlend = len(md.get("missingPositions") or [])

    eindeutig = len(dl) == 1
    codes = [c for c in codes if c] or ["?"]
    code = codes[0] if eindeutig else "?"
    # Kein Ergebnis: PharmCAT sagt es selbst ueber den Phaenotyp-Text
    kein = (not phen) or (phen and phen[0].strip().lower() in ("no result", "n/a", ""))
    if kein:
        code = "?"
        # Ohne Ergebnis darf auch kein Diplotyp stehen. PharmCAT schreibt dort
        # 'Unknown/Unknown', das saehe in der Oberflaeche wie ein Befund aus.
        a1 = a2 = f1 = f2 = ""
        labels = []
    lvl = LVLOF.get(code, -1)
    kind = KIND.get(g, "marker")
    if kind == "trans": de = DE_TRANS.get(code, "")
    elif g in DE_SPEZIAL: de = DE_SPEZIAL[g].get(code, "")
    else: de = DE_ENZ.get(code, "")
    # Risiko-, Ziel- und HLA-Gene haben keinen Metabolisierertyp. PharmCAT liefert
    # dort eigene Phaenotyp-Woerter. 'Uncertain Susceptibility' etwa heisst: es
    # wurde KEINE bekannte Risikovariante gefunden - das ist ein Ergebnis und
    # keine Luecke. Solche Gene duerfen nicht als 'nicht bestimmbar' gelten.
    ohne_skala = 0
    if not de:
        SPEZ = {"uncertain susceptibility":"Keine Risikovariante gefunden",
                "malignant hyperthermia susceptibility":"Erh\u00f6htes Risiko (maligne Hyperthermie)",
                "ivacaftor non-responsive in cf patients":"Spricht nicht auf Ivacaftor an",
                "normal":"Normale Funktion","no result":"","n/a":""}
        de = SPEZ.get((phen[0] if phen else "").strip().lower(), "")
        if de and eindeutig and not kein and kind in ("risiko", "ziel", "hla", "marker"):
            lvl = 2          # bestimmt, aber ohne Metabolisierer-Skala
            ohne_skala = 1

    genes.append(dict(
        g=g, kind=kind, rolle=esc(ROLLE.get(g, "")),
        dip=esc(labels[0]) if eindeutig and labels else "",
        kand=len(dl) if not eindeutig else 1,
        phen=esc(phen[0]) if phen and not kein else "",
        de=esc(de), code=code, lvl=lvl, score=score, flach=ohne_skala,
        a1=esc(a1), f1=esc(f1), a2=esc(a2), f2=esc(f2),
        ok=1 if (eindeutig and not kein) else 0,
        mehr=1 if not eindeutig else 0,
        pos=da, fehlt=fehlend,
        alt=[esc(x) for x in sorted({d.get("label") for d in dl if d.get("label")})][:12] if not eindeutig else [],
        # alle gerufenen Positionen: rsID + Genotyp, fuer den Arztbericht
        var=[[esc(v.get("dbSnpId") or v.get("position")), esc(v.get("call") or "")]
             for v in (r.get("variants") or []) if v.get("call")],
        # Allele, die durch die gerufenen Positionen unterscheidbar sind. Das ist
        # die einzige belastbare Herkunft fuer "im Test untersuchte Varianten" -
        # eine vollstaendige Allel-Definitionsliste liefert PharmCAT hier nicht.
        alle=sorted({esc(a) for v in (r.get("variants") or [])
                     for a in (v.get("alleles") or []) if a},
                    key=lambda s: (0 if s.startswith("*") else 1,
                                   int(re.sub(r"\D", "", s) or 0) if s.startswith("*") else 0, s))[:80],
        unc=[esc(x) for x in (r.get("uncalledHaplotypes") or [])][:60],
    ))

meta = ph["matcherMetadata"]
pos_da = sum(g["pos"] for g in genes)
pos_fehlt = sum(g["fehlt"] for g in genes)

# ---------------------------------------------------------------- Spreadsheet
wb = load_workbook(XL, read_only=True, data_only=True)
ws = wb.worksheets[0]
it = ws.iter_rows(values_only=True)
h = list(next(it))
ci = {}
for i, n in enumerate(h):
    if isinstance(n, str) and n.strip() and n.strip() not in ci: ci[n.strip()] = i

recs = []
for row in it:
    if not row[0]: continue
    def v(k):
        i = ci.get(k)
        return str(row[i]).strip() if i is not None and i < len(row) and row[i] not in (None, "") else ""
    # Die unbenannten Spalten rechts enthalten den Template-Code. Daraus die
    # Ampelfarbe ziehen, die Novogenia selbst fuer diesen Genotyp vergibt.
    tail = " ".join(str(x) for x in row[15:] if x)
    farbe = ""
    if "BG_COLOR_DRUG_RED" in tail: farbe = "crit"
    elif "BG_COLOR_DRUG_YELLOW" in tail: farbe = "warn"
    elif "BG_COLOR_DRUG_GREEN" in tail: farbe = "ok"
    cond = []
    g1, gt1 = v("GENE(1)"), v("GENOTYPE (1)")
    g2, gt2 = v("GENE (2)"), v("METABOLIZER (2)")
    if g1 and gt1: cond.append([g1, gt1])
    if g2 and gt2: cond.append([g2, gt2])
    gl = [lab for lab, col in (("CPIC","CPIC"),("DPWG","DPWG"),("CPNDS","CPNDS"),("FDA","OTHER")) if v(col)]
    # Warfarin: GENE(1) leer, GENOTYPE(1) ist ein Dosisbereich in mg/Tag.
    # Das ist eine Formel aus CYP2C9 + VKORC1, keine Phaenotyp-Zeile.
    formel = v("FORUMLA VALUE")
    recs.append(dict(drug=str(row[0]).strip(), om=v("OM ID"), cond=cond,
                     txt=esc(v("RECOMMEDNATION")), gl=gl, dose=v("DOSE OVERRIDE"),
                     farbe=farbe, dosis=gt1 if (not g1 and gt1) else "", formel=formel))
wb.close()

# Empfehlungstexte deduplizieren
txts = sorted({r["txt"] for r in recs if r["txt"]})
ti = {t: i for i, t in enumerate(txts)}
GL = ["CPIC", "DPWG", "CPNDS", "FDA"]

# Bedingung in Kanon-Codes uebersetzen. Was das Spreadsheet nicht kennt, bleibt roh.
def cond_codes(cond):
    out = []
    for gen, gt in cond:
        c = SHEET.get(gt.upper(), "")
        out.append([gen, c, esc(gt)])       # Code leer = Diplotyp-/Sondervergleich
    return out

rectab = []
for r in recs:
    rectab.append([
        esc(r["drug"]), r["om"], cond_codes(r["cond"]),
        ti.get(r["txt"], -1),
        sum(1 << GL.index(x) for x in r["gl"] if x in GL),
        r["dose"], r["farbe"], esc(r["dosis"]),
    ])

# ---------------------------------------------------------------- schreiben
with open(OUT, "w", encoding="ascii", newline="\n") as f:
    f.write("/* ============================================================\n")
    f.write("   GENE-IUS PGx - echtes Genprofil aus den PharmCAT-Ausgaben\n")
    f.write("   Erzeugt von build_pharmcat.py. Nicht von Hand aendern.\n")
    f.write("   Probe %s | %s | PharmCAT %s\n" % (meta.get("sampleId"), meta.get("genomeBuild"),
                                                  meta.get("namedAlleleMatcherVersion")))
    f.write("   Gene: %d | Positionen gerufen: %d | erwartet aber fehlend: %d\n"
            % (len(genes), pos_da, pos_fehlt))
    f.write("   Leitlinienzeilen: %d fuer %d Wirkstoffe\n"
            % (len(rectab), len({r['drug'] for r in recs})))
    f.write("   ============================================================ */\n")
    f.write("const P_META=%s;\n" % j(dict(
        probe=esc(meta.get("sampleId")), build=esc(meta.get("genomeBuild")),
        ver=esc(meta.get("namedAlleleMatcherVersion")),
        stand=esc(str(meta.get("timestamp"))[:10]),
        posda=pos_da, posfehlt=pos_fehlt,
        quelle=esc(meta.get("inputFilename")))))
    f.write("/* je Gen: Symbol, Art, Rolle, Diplotyp, Phaenotyp, deutsch, Code, Stufe,\n")
    f.write("   Aktivitaetsscore, Allel1+Funktion, Allel2+Funktion, eindeutig, mehrdeutig,\n")
    f.write("   Positionen gerufen/fehlend, Kandidaten, gerufene Varianten, nicht rufbare Allele */\n")
    f.write("const P_GENES=%s;\n" % j(genes))
    f.write("const P_TXT=%s;\n" % j(txts))
    f.write("const P_GL=%s;\n" % j(GL))
    f.write("/* [Wirkstoff, OM-ID, [[Gen,Code,Rohwert],...], Textindex, Leitlinienmaske,\n")
    f.write("    DoseOverride, Ampelfarbe aus dem Template, Dosisbereich (nur Warfarin)] */\n")
    f.write("const P_REC=%s;\n" % j(rectab))

print("geschrieben: %s  %.1f kB" % (OUT, os.path.getsize(OUT) / 1024.0))
print("\nGene: %d | eindeutig gerufen: %d | mehrdeutig: %d | kein Ergebnis: %d"
      % (len(genes), sum(g["ok"] for g in genes), sum(g["mehr"] for g in genes),
         sum(1 for g in genes if not g["ok"] and not g["mehr"])))
print("Positionen: %d gerufen, %d fehlend (%.1f %% Abdeckung)"
      % (pos_da, pos_fehlt, 100.0 * pos_da / (pos_da + pos_fehlt)))
print("Leitlinienzeilen: %d | Wirkstoffe: %d | mit Ampelfarbe: %d"
      % (len(rectab), len({r['drug'] for r in recs}), sum(1 for r in recs if r["farbe"])))
c = Counter(r["farbe"] or "keine" for r in recs)
print("  Farbverteilung:", dict(c))

# Kontrollausgabe: welche Zeilen treffen auf dieses Profil zu?
prof = {g["g"]: g for g in genes}
def trifft(cond):
    """(True/False/None, Grund) - None heisst nicht entscheidbar."""
    if not cond: return (None, "keine Genbedingung (Dosisformel)")
    unklar = None
    for gen, code, roh in cond:
        p = prof.get(gen)
        if not p: return (None, "Gen %s nicht im Panel" % gen)
        if p["mehr"]: unklar = "%s mehrdeutig (%d Kandidaten)" % (gen, p["kand"]); continue
        if not p["ok"]: unklar = "%s ohne Ergebnis" % gen; continue
        mine = {p["code"]}
        if p["code"] in AEQ: mine.add(AEQ[p["code"]])
        if code:
            if code not in mine: return (False, "%s ist %s, verlangt %s" % (gen, p["code"], code))
        else:
            if p["dip"].replace(" ", "").upper() != roh.replace(" ", "").upper():
                return (False, "%s Diplotyp %s != %s" % (gen, p["dip"], roh))
    return (None, unklar) if unklar else (True, "alle Bedingungen erfuellt")

zae = Counter(); tr = []
for r in recs:
    ok, warum = trifft(cond_codes(r["cond"]))
    zae["trifft zu" if ok is True else ("trifft nicht zu" if ok is False else "nicht entscheidbar")] += 1
    if ok is True: tr.append((r["drug"], r["farbe"], r["txt"][:80]))
print("\nAbgleich gegen dieses Profil:", dict(zae))
for x in tr: print("   TREFFER %-18s %-5s %s" % x)
