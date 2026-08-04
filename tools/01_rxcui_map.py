# -*- coding: utf-8 -*-
"""
Schritt 1: Alle 2.694 Wirkstoffe aus All Drugs V12 auf RxNorm-Konzepte (RXCUI) abbilden.
Das ist die Grundlage fuer alle drei Datensaetze (Alternativen, Handelsnamen, Interaktionen).

Zweistufig wie in der Recherche empfohlen:
  1. exakte Suche  /REST/rxcui.json?name=X&search=2
  2. Naeherung     /REST/approximateTerm.json  (nur bei Fehlschlag)

Zusaetzlich wird die ClinPGx-drugs.tsv als dritte Quelle herangezogen - dort stehen
RxNorm-IDs bereits fertig drin, das spart Treffer und faengt Schreibweisen ab.

Ausgabe: rxcui_map.json  { wirkstoff_lower: {name, rxcui, quelle, score} }
"""
import csv, io, json, os, re, sys, time, urllib.parse, urllib.request, urllib.error, zipfile
from openpyxl import load_workbook

csv.field_size_limit(10**7)
B = r"C:\Users\DanielWallerstorfer\Novogenia GmbH\AI RESOURCES - Dokumente\PHARMACOGENETICS"
UA = {"User-Agent": "Novogenia-PGx-Pipeline/1.0 (kontakt: office@novogenia.com)"}
OUT = "rxcui_map.json"

def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower() if s else ""

def get(url, tries=3):
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers=UA)
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.load(r)
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep(2 + i * 2); continue
            return None
        except Exception:
            time.sleep(0.5 + i); continue
    return None

# ---------------------------------------------------------------- Wirkstoffe
wb = load_workbook(B + r"\All Drugs V12.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[0]; it = ws.iter_rows(values_only=True); h = list(next(it))
iName = h.index("DRUG_NAME")
namen = []
for r in it:
    if r[iName]:
        namen.append(str(r[iName]).strip())
wb.close()
print("Wirkstoffe aus All Drugs V12: %d" % len(namen))

# ------------------------------------------------- Vorbelegung aus ClinPGx
print("lade ClinPGx drugs.tsv als Vorbelegung ...")
vor = {}
try:
    req = urllib.request.Request(
        "https://api.clinpgx.org/v1/download/file/data/drugs.zip", headers=UA)
    with urllib.request.urlopen(req, timeout=180) as r:
        z = zipfile.ZipFile(io.BytesIO(r.read()))
    for row in csv.DictReader(
            io.StringIO(z.read("drugs.tsv").decode("utf-8", "replace")), delimiter="\t"):
        rx = (row.get("RxNorm Identifiers") or "").strip()
        if not rx:
            continue
        rx = rx.split(",")[0].strip()
        for n in [row["Name"]] + [x.strip() for x in (row.get("Generic Names") or "").split(",")]:
            if n:
                vor.setdefault(norm(n), rx)
    print("  %d Namensvarianten mit RxNorm-ID vorbelegt" % len(vor))
except Exception as e:
    print("  ClinPGx nicht erreichbar (%s) - weiter ohne Vorbelegung" % e)

# ------------------------------------------------------------------ Mapping
mapping = {}
if os.path.exists(OUT):                       # Wiederaufnahme nach Abbruch
    mapping = json.load(open(OUT, encoding="utf-8"))
    print("  %d Eintraege aus vorherigem Lauf uebernommen" % len(mapping))

stat = {"clinpgx": 0, "exakt": 0, "naeherung": 0, "kein_treffer": 0}
t0 = time.time()
for i, name in enumerate(namen):
    k = norm(name)
    if k in mapping:
        stat[mapping[k]["quelle"]] = stat.get(mapping[k]["quelle"], 0) + 1
        continue

    if k in vor:                                                  # 1) ClinPGx
        mapping[k] = {"name": name, "rxcui": vor[k], "quelle": "clinpgx", "score": 100}
        stat["clinpgx"] += 1
    else:
        d = get("https://rxnav.nlm.nih.gov/REST/rxcui.json?name=%s&search=2"
                % urllib.parse.quote(name))
        ids = (d or {}).get("idGroup", {}).get("rxnormId", [])
        if ids:                                                   # 2) exakt
            mapping[k] = {"name": name, "rxcui": ids[0], "quelle": "exakt", "score": 100}
            stat["exakt"] += 1
        else:                                                     # 3) Naeherung
            d = get("https://rxnav.nlm.nih.gov/REST/approximateTerm.json?term=%s&maxEntries=1"
                    % urllib.parse.quote(name))
            c = (d or {}).get("approximateGroup", {}).get("candidate", [])
            if c and c[0].get("rxcui"):
                sc = float(c[0].get("score", 0))
                mapping[k] = {"name": name, "rxcui": c[0]["rxcui"],
                              "quelle": "naeherung", "score": sc}
                stat["naeherung"] += 1
            else:
                mapping[k] = {"name": name, "rxcui": None,
                              "quelle": "kein_treffer", "score": 0}
                stat["kein_treffer"] += 1
        time.sleep(0.06)                       # ~16 req/s, Limit ist 20

    if (i + 1) % 200 == 0:
        json.dump(mapping, open(OUT, "w", encoding="utf-8"), ensure_ascii=False)
        print("  %4d/%d  %.0fs  %s" % (i + 1, len(namen), time.time() - t0, stat))

json.dump(mapping, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
mit = sum(1 for v in mapping.values() if v["rxcui"])
print("\nfertig in %.0fs" % (time.time() - t0))
print("Wirkstoffe mit RXCUI: %d von %d = %.0f %%" % (mit, len(mapping), 100 * mit / len(mapping)))
print("Quellen:", stat)
schwach = [v for v in mapping.values() if v["quelle"] == "naeherung" and v["score"] < 50]
print("Naeherungstreffer mit schwachem Score (<50), zur Handpruefung: %d" % len(schwach))
for v in schwach[:10]:
    print("   %-34s -> RXCUI %s (Score %.1f)" % (v["name"][:34], v["rxcui"], v["score"]))
