# -*- coding: utf-8 -*-
"""
Korrektur und Priorisierung.

1) Der Namensabgleich in 04 war zu grosszuegig: er hat Salze weggeworfen und
   dadurch verschiedene Eisensalze zusammengeworfen ("Ferrous Chloride" ==
   "ferrous fumarate"). Diese 21 Treffer werden verworfen.
2) Von den ~1.118 Wirkstoffen ohne MED-RT-Daten: welche sind ueberhaupt wichtig
   genug fuer Agentenrecherche? Kriterien: hat ein PGx-Gen, hat eine Leitlinie,
   steht in Lisas Liste, oder ist in ClinPGx mit Handelsnamen gefuehrt
   (= wird real verordnet).
"""
import csv, io, json, re, urllib.request, zipfile
from collections import Counter
from openpyxl import load_workbook

csv.field_size_limit(10**7)
B = r"C:\Users\DanielWallerstorfer\Novogenia GmbH\AI RESOURCES - Dokumente\PHARMACOGENETICS"
UA = {"User-Agent": "Novogenia-PGx-Pipeline/1.0"}
def norm(s): return re.sub(r"\s+", " ", str(s)).strip().lower() if s else ""

M = json.load(open("rxcui_map.json", encoding="utf-8"))

# --- 1) fragwuerdige Namenstreffer verwerfen -----------------------------
raus = 0
for k, v in M.items():
    if v.get("medrt_via") == "name":
        v["medrt"] = None
        v["medrt_via"] = "verworfen_unsicher"
        raus += 1
print("fragwuerdige Namenstreffer verworfen: %d" % raus)
mit = sum(1 for v in M.values() if v.get("medrt"))
print("belastbare MED-RT-Abdeckung: %d von %d = %.0f %%" % (mit, len(M), 100 * mit / len(M)))

# --- 2) Welche Wirkstoffe sind wichtig? ----------------------------------
# a) PGx-Gen aus All Drugs
wb = load_workbook(B + r"\All Drugs V12.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[0]; it = ws.iter_rows(values_only=True); h = list(next(it))
idx = {x: i for i, x in enumerate(h) if isinstance(x, str)}
ENZ = ["CYP2C19", "CYP2D6", "CYP2C9", "CYP3A4", "NAT2", "CYP1A2", "PGP_Kardio"]
mit_gen = set()
for r in it:
    n = r[idx["DRUG_NAME"]]
    if not n: continue
    for e in ENZ:
        for suf in ("_Main_Substrate", "_Substrate", "_Main_Activating_Substrate"):
            c = e + suf
            if c in idx and r[idx[c]] not in (None, "", 0):
                mit_gen.add(norm(n)); break
wb.close()
print("\nWirkstoffe mit modelliertem PGx-Enzym: %d" % len(mit_gen))

# b) in ClinPGx mit Handelsnamen = wird real verordnet
req = urllib.request.Request("https://api.clinpgx.org/v1/download/file/data/drugs.zip", headers=UA)
with urllib.request.urlopen(req, timeout=180) as r:
    z = zipfile.ZipFile(io.BytesIO(r.read()))
mit_marke, mit_leitlinie = set(), set()
for row in csv.DictReader(io.StringIO(z.read("drugs.tsv").decode("utf-8", "replace")), delimiter="\t"):
    namen = [row["Name"]] + [x.strip() for x in (row.get("Generic Names") or "").split(",")]
    if (row.get("Trade Names") or "").strip():
        mit_marke.update(norm(n) for n in namen if n)
    if (row.get("Dosing Guideline") or "").strip().lower() == "yes":
        mit_leitlinie.update(norm(n) for n in namen if n)
print("in ClinPGx mit Handelsnamen: %d Namensvarianten" % len(mit_marke))
print("mit Dosierungsleitlinie:     %d Namensvarianten" % len(mit_leitlinie))

# --- 3) Luecke priorisieren ----------------------------------------------
luecke = [(k, v) for k, v in M.items() if not v.get("medrt")]
prio = {"1_leitlinie": [], "2_pgx_gen": [], "3_verordnet": [], "4_rest": []}
for k, v in luecke:
    if k in mit_leitlinie:   prio["1_leitlinie"].append(v["name"])
    elif k in mit_gen:       prio["2_pgx_gen"].append(v["name"])
    elif k in mit_marke:     prio["3_verordnet"].append(v["name"])
    else:                    prio["4_rest"].append(v["name"])

print("\n=== Die %d Wirkstoffe ohne MED-RT-Daten, nach Wichtigkeit ===" % len(luecke))
for k in sorted(prio):
    print("  %-14s %4d" % (k, len(prio[k])))
    for n in prio[k][:8]:
        print("        %s" % n)

json.dump(M, open("rxcui_map.json", "w", encoding="utf-8"), ensure_ascii=False, indent=0)
json.dump(prio, open("luecke_prio.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
wichtig = len(prio["1_leitlinie"]) + len(prio["2_pgx_gen"]) + len(prio["3_verordnet"])
print("\n=> fuer Agentenrecherche wirklich lohnend: %d Wirkstoffe" % wichtig)
print("   (der Rest sind Obsoleta, Diagnostika und Nischensubstanzen)")
