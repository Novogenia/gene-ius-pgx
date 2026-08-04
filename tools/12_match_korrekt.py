# -*- coding: utf-8 -*-
"""
Korrigierter Abgleich PharmCAT <-> Novogenia-Spreadsheets.

Drei Korrekturen gegenueber dem ersten Versuch:
  a) 'phenotypes' hat Vorrang vor 'lookupKey'. Bei CYP2C9 und DPYD ist der
     lookupKey der Aktivitaetsscore ("2.0"), der Phaenotyp steht daneben.
  b) Zusaetzlich zum Phaenotyp wird auch das Diplotyp-Label verglichen -
     das Spreadsheet hat Zeilen wie '*28/*28' (UGT1A1) und 'TT' (VKORC1).
  c) Kein Treffer bei NORMAL ist KEIN Fehlschlag, sondern die Aussage
     "keine Anpassung notwendig". Nur ein fehlender Treffer bei einem
     auffaelligen Phaenotyp ist eine Luecke im Spreadsheet.
"""
import csv, json, os, re
from collections import defaultdict, Counter
from openpyxl import load_workbook

B = r"C:\Users\DanielWallerstorfer\Novogenia GmbH\AI RESOURCES - Dokumente\PHARMACOGENETICS"
D = r"C:\Users\DanielWallerstorfer\Downloads"
def kopf(t): print("\n" + "=" * 74 + "\n" + t + "\n" + "=" * 74)

# --------------------------------------------------- PharmCAT auslesen
pc = json.load(open(os.path.join(D, "pharmcat.phenotype (1).json"), encoding="utf-8"))
gr = pc["geneReports"]

# Phaenotyp -> Kanon. Deckt CPIC-Metabolisierer und Transporter-Funktion ab.
CANON = {
 "ultrarapid metabolizer":"UM","rapid metabolizer":"RM","normal metabolizer":"NM",
 "intermediate metabolizer":"IM","poor metabolizer":"PM","likely poor metabolizer":"PM",
 "likely intermediate metabolizer":"IM",
 "normal function":"NF","increased function":"IF","decreased function":"DF",
 "poor function":"PF","possible decreased function":"DF",
 "normal":"NM","indeterminate":"?","no result":"?","n/a":"?",
}
# Spreadsheet-Vokabular -> derselbe Kanon
SHEET = {"ULTRARAPID":"UM","RAPID":"RM","EXTENSIVE":"NM","NORMAL":"NM",
         "INTERMEDIATE":"IM","POOR":"PM"}
# Transporter/Enzym-Aequivalenz: das Spreadsheet nutzt fuer SLCO1B1 die
# Metabolisierer-Woerter, PharmCAT die Funktions-Woerter.
AEQUIV = {"NF":"NM","DF":"IM","PF":"PM","IF":"UM"}

def canon_pc(s):
    s = (s or "").strip().lower()
    return CANON.get(s, "")

def canon_sheet(s):
    s = (s or "").strip().upper()
    return SHEET.get(s, s)

profil = {}     # Gen -> dict(code, codes, label, labels, mehrdeutig, phen)
for g, r in gr.items():
    dl = r.get("recommendationDiplotypes") or []
    codes, labels, phen = [], [], []
    for d in dl:
        for p in (d.get("phenotypes") or []):
            c = canon_pc(p)
            if c and c not in codes: codes.append(c)
            if p not in phen: phen.append(p)
        if d.get("label"): labels.append(d["label"])
    # Wenn kein Phaenotyp da ist, den lookupKey nehmen (z.B. reine Scores)
    lk = sorted({k for d in dl for k in (d.get("lookupKey") or [])})
    profil[g] = dict(codes=codes, labels=labels, phen=phen, lookup=lk,
                     mehrdeutig=len(dl) > 1,
                     score=sorted({str(d.get("activityScore")) for d in dl
                                   if d.get("activityScore") not in (None,"","n/a")}),
                     uncalled=r.get("uncalledHaplotypes") or [])

kopf("Lisas Genprofil laut PharmCAT (23 Gene)")
print("%-9s %-24s %-13s %-6s %s" % ("GEN","Phaenotyp","Diplotyp","Code","Hinweis"))
print("-"*74)
for g in sorted(profil):
    v = profil[g]
    code = v["codes"][0] if len(v["codes"]) == 1 else ("/".join(v["codes"]) or "?")
    hint = ""
    if v["mehrdeutig"]: hint = "%d moegliche Diplotypen" % len(v["labels"])
    elif not v["codes"]: hint = "nicht gerufen (Positionen fehlen)"
    if v["score"]: hint = ("AS=%s " % ",".join(v["score"])) + hint
    print("%-9s %-24s %-13s %-6s %s" % (
        g, (v["phen"][0] if v["phen"] else (v["lookup"][0] if v["lookup"] else "-"))[:24],
        (v["labels"][0] if v["labels"] else "-")[:13], code, hint))

# --------------------------------------------------- Spreadsheet lesen
wb = load_workbook(B + r"\Pharmgkb drug recommendations V4.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[0]
it = ws.iter_rows(values_only=True)
h = list(next(it))
ci = {}
for i, n in enumerate(h):
    if isinstance(n, str) and n.strip() not in ci: ci[n.strip()] = i
rows = []
for row in it:
    dn = row[ci["DRUG"]]
    if not dn: continue
    def val(k):
        i = ci.get(k)
        return str(row[i]).strip() if i is not None and row[i] not in (None,"") else ""
    rows.append(dict(drug=str(dn).strip(), g1=val("GENE(1)"), gt1=val("GENOTYPE (1)"),
                     g2=val("GENE (2)"), gt2=val("METABOLIZER (2)"),
                     rec=val("RECOMMEDNATION"),
                     cpic=val("CPIC"), dpwg=val("DPWG"), other=val("OTHER"),
                     dose=val("DOSE OVERRIDE")))
wb.close()

# --------------------------------------------------- Abgleich je Zeile
def passt(gen, sheet_gt):
    """Trifft diese Spreadsheet-Zeile auf Lisas Profil zu?"""
    v = profil.get(gen)
    if not v: return (None, "Gen nicht in PharmCAT-Panel")
    if v["mehrdeutig"]:
        cs = set(v["codes"])
        c = canon_sheet(sheet_gt)
        if c in cs: return (None, "mehrdeutig - Treffer moeglich, nicht sicher")
        return (None, "mehrdeutig - nicht entscheidbar")
    if not v["codes"] and not v["labels"]:
        return (None, "kein Ergebnis - Positionen fehlen im VCF")
    c = canon_sheet(sheet_gt)
    mine = set(v["codes"]) | {AEQUIV[x] for x in v["codes"] if x in AEQUIV}
    if c in mine: return (True, "Phaenotyp-Treffer")
    # Diplotyp-Ebene: '*28/*28' oder 'TT'
    for lab in v["labels"]:
        if lab.replace(" ", "").upper() == sheet_gt.replace(" ", "").upper():
            return (True, "Diplotyp-Treffer")
    return (False, "trifft nicht zu")

kopf("Abgleich der 103 Empfehlungszeilen gegen Lisas Profil")
zaehler = Counter(); treffer = []; unklar = defaultdict(list)
for r in rows:
    if not r["g1"]: continue
    ok, warum = passt(r["g1"], r["gt1"])
    if ok is True:
        zaehler["trifft zu"] += 1; treffer.append(r)
    elif ok is False:
        zaehler["trifft nicht zu"] += 1
    else:
        zaehler[warum] += 1; unklar[warum].append("%s (%s %s)" % (r["drug"], r["g1"], r["gt1"]))
for k, v in zaehler.most_common(): print("  %-46s %3d" % (k, v))

kopf("Diese Wirkstoffe bekommen eine genotypspezifische Empfehlung")
for r in treffer:
    gl = ",".join(x for x, y in (("CPIC", r["cpic"]), ("DPWG", r["dpwg"]), ("FDA", r["other"])) if y)
    print("  %-22s %s %-13s [%s]" % (r["drug"], r["g1"], r["gt1"], gl or "-"))
    print("      %s" % r["rec"][:190])
    if r["dose"]: print("      Dosis-Override: %s" % r["dose"])

kopf("Was nicht entscheidbar ist - und warum")
for warum, lst in unklar.items():
    print("\n%s  (%d Zeilen)" % (warum.upper(), len(lst)))
    for x in sorted(set(lst))[:26]: print("    " + x)
    if len(set(lst)) > 26: print("    ... und %d weitere" % (len(set(lst)) - 26))

# --------------------------------------------------- Luecken im Spreadsheet
kopf("Luecken: auffaelliger Phaenotyp, aber keine Zeile im Spreadsheet")
have = defaultdict(set)
for r in rows:
    if r["g1"]: have[r["g1"]].add(canon_sheet(r["gt1"]))
for g in sorted(profil):
    v = profil[g]
    if v["mehrdeutig"] or not v["codes"]: continue
    c = v["codes"][0]
    if c in ("NM", "NF", "?"): continue          # normal = kein Handlungsbedarf
    if g not in have:
        print("  %-9s %-6s kein einziger Wirkstoff im Spreadsheet auf dieses Gen bezogen" % (g, c))
    elif c not in have[g] and not (c in AEQUIV and AEQUIV[c] in have[g]):
        print("  %-9s %-6s Spreadsheet hat nur %s -> Empfehlung fehlt" % (g, c, sorted(have[g])))

# --------------------------------------------------- App-Gene vs PharmCAT
kopf("Gene der App vs. PharmCAT-Panel")
APP = ["CYP2C19","CYP2D6","CYP2C9","VKORC1","DPYD","SLCO1B1","TPMT","UGT1A1",
       "CYP3A4","NAT2","ABCB1","G6PD","CYP1A2"]
for g in APP:
    v = profil.get(g)
    if not v: print("  %-9s NICHT im PharmCAT-Panel" % g)
    elif v["mehrdeutig"]: print("  %-9s mehrdeutig (%d Diplotypen)" % (g, len(v["labels"])))
    elif v["codes"]: print("  %-9s %s (%s)" % (g, v["codes"][0], v["labels"][0] if v["labels"] else "-"))
    else: print("  %-9s kein Ergebnis" % g)
neu = sorted(set(profil) - set(APP))
print("\nZusaetzlich von PharmCAT geliefert (in der App noch nicht modelliert):")
print("  " + ", ".join(neu))
