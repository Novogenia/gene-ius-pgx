# -*- coding: utf-8 -*-
"""
Baut aus den Novogenia-Quelldaten einen kompakten JS-Datenblock fuer den
GENE-IUS PGx Clickdummy.

Quellen (alle unter AI RESOURCES - Dokumente/PHARMACOGENETICS):
  All Drugs V12.xlsx                    Wirkstoffe, ATC-Ebenen, Enzym-Rollen
  Pharmgkb drug recommendations V4.xlsx offizielle Empfehlungstexte + Leitlinien
  drugs_master.csv                      bestes Evidenzniveau je Wirkstoff
  drug_pharmacogenetics.csv             PGx-Annotationen (Gene, Effekte)
  drug_interactions.csv                 Freitext zu Wechselwirkungen

Regel: nichts erfinden. Fehlt ein Wert, bleibt das Feld leer.
"""
import csv, json, re, sys
from collections import defaultdict, Counter
from openpyxl import load_workbook

B = r"C:\Users\DanielWallerstorfer\Novogenia GmbH\AI RESOURCES - Dokumente\PHARMACOGENETICS"
OUT = "pgx_data.js"

# Gene, fuer die die App einen Phaenotyp kennt
MODELLED = ["CYP2C19","CYP2D6","CYP2C9","VKORC1","DPYD","SLCO1B1","TPMT",
            "UGT1A1","CYP3A4","NAT2","ABCB1","G6PD","CYP1A2"]
# Spaltenname in All Drugs -> Genname der App
ENZ2GENE = {"CYP2C19":"CYP2C19","CYP2D6":"CYP2D6","CYP2C9":"CYP2C9","CYP3A4":"CYP3A4",
            "NAT2":"NAT2","CYP1A2":"CYP1A2","PGP_Kardio":"ABCB1"}

ATC_DE = {
 "ALIMENTARY TRACT AND METABOLISM":"Verdauung &amp; Stoffwechsel",
 "BLOOD AND BLOOD FORMING ORGANS":"Blut",
 "CARDIOVASCULAR SYSTEM":"Herz &amp; Kreislauf",
 "DERMATOLOGICALS":"Haut",
 "GENITO URINARY SYSTEM AND SEX HORMONES":"Urogenital &amp; Sexualhormone",
 "SYSTEMIC HORMONAL PREPARATIONS, EXCL. SEX HORMONES AND INSULINS":"Hormone",
 "ANTIINFECTIVES FOR SYSTEMIC USE":"Infektionen",
 "ANTINEOPLASTIC AND IMMUNOMODULATING AGENTS":"Krebs &amp; Immunsystem",
 "MUSCULO-SKELETAL SYSTEM":"Muskeln &amp; Skelett",
 "NERVOUS SYSTEM":"Nervensystem",
 "ANTIPARASITIC PRODUCTS, INSECTICIDES AND REPELLENTS":"Parasiten",
 "RESPIRATORY SYSTEM":"Atemwege",
 "SENSORY ORGANS":"Sinnesorgane",
 "VARIOUS":"Verschiedenes",
}

def norm(s):
    return re.sub(r"\s+"," ",str(s)).strip().lower() if s is not None else ""

def title(s):
    s = str(s).strip()
    return s[0].upper()+s[1:] if s else s

def esc(s):
    """ASCII-sicher: Umlaute und Sonderzeichen als HTML-Entities."""
    if s is None: return ""
    s = re.sub(r"\s+"," ",str(s)).strip()
    rep = {"ä":"&auml;","ö":"&ouml;","ü":"&uuml;","Ä":"&Auml;","Ö":"&Ouml;","Ü":"&Uuml;",
           "ß":"&szlig;","–":"&mdash;","—":"&mdash;","’":"&rsquo;","‘":"&lsquo;",
           "“":"&ldquo;","”":"&ldquo;","„":"&bdquo;","≥":"&ge;","≤":"&le;","→":"&rarr;",
           "×":"x","·":"&middot;","é":"&eacute;","è":"&egrave;","ï":"&iuml;","°":"&deg;",
           "\u00a0":" ","\u2011":"-","\u2013":"&mdash;","\u2192":"&rarr;"}
    for a,b in rep.items(): s = s.replace(a,b)
    s = s.replace("\\","").replace("`","'").replace("${","$ {")
    return "".join(c if ord(c)<128 else "" for c in s)

# ---------------------------------------------------------------- All Drugs
print("lese All Drugs V12 ...")
wb = load_workbook(B+r"\All Drugs V12.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[0]
it = ws.iter_rows(values_only=True)
hdr = list(next(it))
idx = {h:i for i,h in enumerate(hdr) if isinstance(h,str)}
ENZ = sorted({h[:-len("_Main_Substrate")] for h in hdr
              if isinstance(h,str) and h.endswith("_Main_Substrate")})

def flag(row, col):
    v = row[idx[col]] if col in idx else None
    return v not in (None,"",0)

drugs = {}          # key -> dict
order = []
atc4 = defaultdict(list)       # ATC-Ebene 4 -> [key] (gleiche Indikationsgruppe)
inhibits = defaultdict(list)   # enzym -> [key]  (Hemmer)
induces  = defaultdict(list)
substrate= defaultdict(list)   # enzym -> [key]  (Hauptsubstrat)

for row in it:
    name = row[idx["DRUG_NAME"]]
    if not name: continue
    key = norm(name)
    if key in drugs: continue
    l1 = row[idx["LEVEL_1"]]; l2 = row[idx["LEVEL_2"]]
    if l1 == "#VALUE!": l1 = None
    main_g, sub_g, pro_g = None, None, None
    genes = []
    for e in ENZ:
        g = ENZ2GENE.get(e)
        if flag(row, e+"_Main_Activating_Substrate") and g and not pro_g: pro_g = g
        if flag(row, e+"_Main_Substrate"):
            if g and not main_g: main_g = g
            if g: genes.append(g)
        elif flag(row, e+"_Substrate"):
            if g and not sub_g: sub_g = g
            if g: genes.append(g)
        if flag(row, e+"_Main_Inhibotor") or flag(row, e+"_Inhibitor"):
            inhibits[e].append(key)
        if flag(row, e+"_Main_Inducer") or flag(row, e+"_Inducer"):
            induces[e].append(key)
        if flag(row, e+"_Main_Substrate"):
            substrate[e].append(key)
    l3 = row[idx["LEVEL_3"]]; l4 = row[idx["LEVEL_4"]]
    if l3 == "#VALUE!": l3 = None
    if l4 == "#VALUE!": l4 = None
    if l4: atc4[l4].append(key)
    drugs[key] = dict(name=title(name), l1=l1, l2=l2, l3=l3, l4=l4,
                      gene=main_g or sub_g, genes=sorted(set(genes)),
                      prodrug=bool(pro_g), rec="", gl=[], ev="", recs=0, om=row[idx["OM ID"]])
    order.append(key)
wb.close()
print("  Wirkstoffe:", len(drugs))

# ------------------------------------------- Pharmgkb Empfehlungen (Novogenia)
print("lese Pharmgkb drug recommendations V4 ...")
wb = load_workbook(B+r"\Pharmgkb drug recommendations V4.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[0]
it = ws.iter_rows(values_only=True)
h = list(next(it))
c = {n:h.index(n) for n in ["DRUG","GENE(1)","GENOTYPE (1)","GENE (2)","METABOLIZER (2)",
                            "RECOMMEDNATION","CPIC","DPWG","CPNDS","OTHER"] if n in h}
# GENOTYPE (1) -> Metabolisierer-Stufe der App (0 langsam ... 3 ultraschnell)
PHENO_CODE = {"POOR":0, "INTERMEDIATE":1, "EXTENSIVE":2, "NORMAL":2, "ULTRARAPID":3}
recrows = []          # (drugkey, Gen, Stufe, Roh-Genotyp, Text, [Leitlinien])
for row in it:
    dn = row[c["DRUG"]] if "DRUG" in c else None
    if not dn: continue
    key = norm(dn)
    d = drugs.get(key)
    if d is None:
        d = dict(name=title(dn), l1=None, l2=None, l3=None, l4=None, gene=None, genes=[], prodrug=False,
                 rec="", gl=[], ev="", recs=0, om=None)
        drugs[key] = d; order.append(key)
    g1 = str(row[c["GENE(1)"]]).strip() if row[c["GENE(1)"]] else None
    g2 = str(row[c["GENE (2)"]]).strip() if row[c["GENE (2)"]] else None
    for g in (g1,g2):
        if g and g in MODELLED:
            if g not in d["genes"]: d["genes"].append(g)
            if not d["gene"] or d["gene"] not in MODELLED: d["gene"] = g
    gl = []
    for lab,col in (("CPIC","CPIC"),("DPWG","DPWG"),("CPNDS","CPNDS"),("FDA","OTHER")):
        if col in c and row[c[col]] not in (None,"",0):
            gl.append(lab)
            if lab not in d["gl"]: d["gl"].append(lab)
    txt = row[c["RECOMMEDNATION"]] if "RECOMMEDNATION" in c else None
    gt = str(row[c["GENOTYPE (1)"]]).strip() if row[c["GENOTYPE (1)"]] else ""
    if txt and g1:
        recrows.append((key, g1, PHENO_CODE.get(gt.upper(), -1), esc(gt), esc(txt), gl))
        if not d["rec"]: d["rec"] = esc(txt)
wb.close()
print("  Empfehlungszeilen:", len(recrows), "fuer", len({r[0] for r in recrows}), "Wirkstoffe")

# ------------------------------------------------ Evidenz aus drugs_master.csv
EVMAP = {"x1a":"guideline","x1b":"label","x2a":"reviewed","x2b":"reviewed",
         "x3":"studies","x4":"studies"}
nev = 0
with open(B+r"\drugs_master.csv", encoding="utf-8-sig", newline="") as f:
    for r in csv.DictReader(f):
        key = norm(r["drug"])
        d = drugs.get(key)
        if not d: continue
        e = (r.get("pgx_best_evidence") or "").strip().lower()
        if e in EVMAP: d["ev"] = EVMAP[e]; nev += 1
        for g in (r.get("pgx_genes") or "").replace(";",",").split(","):
            g = g.strip()
            if g in MODELLED and g not in d["genes"]: d["genes"].append(g)
        if not d["gene"]:
            for g in d["genes"]:
                if g in MODELLED: d["gene"] = g; break
print("  Evidenzstufen:", nev)

# --------------------------------------- Anzahl PGx-Annotationen je Wirkstoff
ann = Counter()
with open(B+r"\drug_pharmacogenetics.csv", encoding="utf-8-sig", newline="") as f:
    for r in csv.DictReader(f):
        ann[norm(r["drug"])] += 1
for k,v in ann.items():
    if k in drugs: drugs[k]["recs"] = v
print("  Wirkstoffe mit Annotationen:", sum(1 for k in drugs if drugs[k]["recs"]))

# ------------------------------------------------ Wechselwirkungen
# Hauptquelle: drug_interactions.csv (DrugBank). Der Satz sagt, was passiert.
print("lese Wechselwirkungen (DrugBank) ...")

def classify(t):
    """Freitext -> (Schweregrad, Wirkungsart) ohne etwas dazuzuerfinden."""
    s = t.lower()
    if "risk or severity" in s and "increased" in s:
        return "crit", "risk"          # Nebenwirkungsrisiko steigt
    if ("serum concentration" in s and "increased" in s) \
       or ("metabolism" in s and "decreased" in s) \
       or ("excretion rate" in s and "higher serum level" in s) \
       or ("excretion of" in s and "decreased" in s):
        return "crit", "up"            # Wirkstoff reichert sich an
    if ("therapeutic efficacy" in s and "decreased" in s) \
       or ("metabolism" in s and "increased" in s) \
       or ("serum concentration" in s and "decreased" in s) \
       or ("excretion rate" in s and "lower serum level" in s):
        return "warn", "down"          # Wirkung faellt schwaecher aus
    return "warn", "other"

EFF_DE = {
  "risk": "Erh&ouml;htes Risiko f&uuml;r Nebenwirkungen",
  "up":   "Wirkstoffspiegel steigt &mdash; Gefahr der Anreicherung",
  "down": "Wirkung f&auml;llt schw&auml;cher aus",
  "other":"Gegenseitige Beeinflussung",
}

dbix = {}
with open(B+r"\drug_interactions.csv", encoding="utf-8-sig", newline="") as f:
    for r in csv.DictReader(f):
        a, b = norm(r["drug_a"]), norm(r["drug_b"])
        if a not in drugs or b not in drugs or a == b: continue
        k = (a,b) if a < b else (b,a)
        if k in dbix: continue
        sev, eff = classify(r["interaction"])
        dbix[k] = (sev, eff, r["interaction"], norm(r["drug_a"]))
print("  Paare, bei denen beide Wirkstoffe bekannt sind:", len(dbix))

# Ergaenzung: enzymatisch abgeleitete Paare, aber nur fuer die Gene, die die App
# auch bewertet - das ist die Verbindung zwischen Interaktion und Genprofil.
enzpairs = {}
for e in ENZ:
    if e not in ENZ2GENE: continue
    subs = set(substrate[e])
    for a in set(inhibits[e]):
        for b in subs:
            if a == b: continue
            k = (a,b) if a < b else (b,a)
            if k in dbix or k in enzpairs: continue
            enzpairs[k] = (e, a)          # a hemmt, b ist Substrat
print("  zusaetzliche enzymatisch abgeleitete Paare:", len(enzpairs))

# Begrenzen: hoechstens 12 Partner je Wirkstoff. Erst DrugBank, dann Enzymlogik.
per = Counter(); ixout = []
for (a,b),(sev,eff,txt,first) in sorted(dbix.items()):
    if per[a] >= 12 or per[b] >= 12: continue
    per[a] += 1; per[b] += 1
    ixout.append((a, b, sev, None, eff, txt, first))
for (a,b),(e,inh) in sorted(enzpairs.items()):
    if per[a] >= 12 or per[b] >= 12: continue
    per[a] += 1; per[b] += 1
    ixout.append((a, b, "warn", e, "enz", "", inh))
print("  aufgenommen:", len(ixout),
      "| davon mit Originaltext:", sum(1 for x in ixout if x[5]))

# ------------------------------------------------ Alternativen ueber ATC-Ebene 4
print("leite Alternativen aus der ATC-Ebene 4 ab ...")
alt_of = defaultdict(list)
for grp, members in atc4.items():
    if len(members) < 2: continue
    for k in members:
        alt_of[k] = [m for m in members if m != k][:40]
print("  Wirkstoffe mit mindestens einer Alternative:", len(alt_of))

# ------------------------------------------------------------ JS schreiben
atc1 = sorted({d["l1"] for d in drugs.values() if d["l1"]})
atc2 = sorted({d["l2"] for d in drugs.values() if d["l2"]})
atc3 = sorted({d.get("l3") for d in drugs.values() if d.get("l3")})
atc4n = sorted({d.get("l4") for d in drugs.values() if d.get("l4")})
a3i = {v:i for i,v in enumerate(atc3)}
a4i = {v:i for i,v in enumerate(atc4n)}
atc3 = sorted({d.get("l3") for d in drugs.values() if d.get("l3")})
atc4n = sorted({d.get("l4") for d in drugs.values() if d.get("l4")})
a3i = {v:i for i,v in enumerate(atc3)}
a4i = {v:i for i,v in enumerate(atc4n)}
a1i = {v:i for i,v in enumerate(atc1)}
a2i = {v:i for i,v in enumerate(atc2)}
genes_used = sorted({d["gene"] for d in drugs.values() if d["gene"]})
gi = {v:i for i,v in enumerate(genes_used)}
texts = sorted({d["rec"] for d in drugs.values() if d["rec"]})
ti = {v:i for i,v in enumerate(texts)}
GL = ["CPIC","DPWG","CPNDS","FDA"]
EV = ["","guideline","label","reviewed","studies"]

keyidx = {}
rows = []
for k in order:
    d = drugs[k]
    keyidx[k] = len(rows)
    glmask = sum(1 << GL.index(g) for g in d["gl"] if g in GL)
    rows.append([
        esc(d["name"]),
        a1i.get(d["l1"], -1),
        a2i.get(d["l2"], -1),
        gi.get(d["gene"], -1),
        1 if d["prodrug"] else 0,
        EV.index(d["ev"]) if d["ev"] in EV else 0,
        glmask,
        ti.get(d["rec"], -1),
        d["recs"],
        [gi[g] for g in d["genes"] if g in gi],
        a3i.get(d.get("l3"), -1),
        a4i.get(d.get("l4"), -1),
    ])

EFFS = ["risk","up","down","other","enz"]
# Die DrugBank-Saetze sind formelhaft. Statt des ganzen Satzes wird nur der
# konkrete Risikobegriff gespeichert ("QTc prolongation", "bleeding", ...).
def risk_subject(t):
    m = re.search(r"risk or severity of ([^,]+?) can be", t, re.I)
    if m:
        s = m.group(1).strip()
        return "" if s.lower() == "adverse effects" else esc(s)[:60]
    m = re.search(r"may (increase|decrease) the (.+?) activities of", t, re.I)
    if m: return esc(m.group(2))[:60]
    return ""
subjects = sorted({risk_subject(t) for _,_,_,_,_,t,_ in ixout if t and risk_subject(t)})
si = {v:i for i,v in enumerate(subjects)}
ix = [[keyidx[a], keyidx[b], {"crit":2,"warn":1}[sev],
       gi.get(ENZ2GENE.get(e or "", ""), -1), EFFS.index(eff),
       si.get(risk_subject(t), -1), 0 if first == a else 1]
      for a,b,sev,e,eff,t,first in ixout if a in keyidx and b in keyidx]
alt = [[keyidx[k]] + [keyidx[m] for m in v if m in keyidx][:12]
       for k,v in alt_of.items() if k in keyidx]

# Genotyp-spezifische Empfehlungen: [Wirkstoff, Gen, Stufe, RohGenotyp, Text, Leitlinienmaske]
rectab = []
for key, gen, code, raw, txt, gl in recrows:
    if key not in keyidx: continue
    rectab.append([keyidx[key], gi.get(gen, -1), code, raw, ti.get(txt, -1),
                   sum(1 << GL.index(g) for g in gl if g in GL)])
# Gennamen, die nicht modelliert sind, trotzdem als Text mitgeben
recgenes = sorted({r[1] for r in recrows})
rgi = {g:i for i,g in enumerate(recgenes)}
for i,(key, gen, code, raw, txt, gl) in enumerate(
        [r for r in recrows if r[0] in keyidx]):
    rectab[i][1] = rgi[gen]

def j(o): return json.dumps(o, separators=(",",":"), ensure_ascii=True)

with open(OUT, "w", encoding="ascii", newline="\n") as f:
    f.write("/* ============================================================\n")
    f.write("   GENE-IUS PGx - Wirkstoffdaten aus den Novogenia-Quelldateien\n")
    f.write("   Erzeugt von build_pgx_data.py. Nicht von Hand aendern.\n")
    f.write("   Quellen: All Drugs V12.xlsx, Pharmgkb drug recommendations V4.xlsx,\n")
    f.write("            drugs_master.csv, drug_pharmacogenetics.csv, drug_interactions.csv\n")
    f.write("   Wirkstoffe: %d | Empfehlungstexte: %d | Wechselwirkungen: %d\n" % (len(rows), len(texts), len(ix)))
    f.write("   ============================================================ */\n")
    f.write("const D_ATC1=%s;\n" % j([esc(x) for x in atc1]))
    f.write("const D_ATC1DE=%s;\n" % j([ATC_DE.get(x, esc(x).title()) for x in atc1]))
    f.write("const D_ATC2=%s;\n" % j([esc(x).title() for x in atc2]))
    f.write("const D_ATC3=%s;\n" % j([esc(x).title() for x in atc3]))
    f.write("const D_ATC4=%s;\n" % j([esc(x).title() for x in atc4n]))
    f.write("const D_GENES=%s;\n" % j(genes_used))
    f.write("const D_GL=%s;\n" % j(GL))
    f.write("const D_EV=%s;\n" % j(EV))
    f.write("const D_TXT=%s;\n" % j(texts))
    f.write("/* [Name, atc1, atc2, Gen, Prodrug, Evidenz, Leitlinienmaske, Textindex, Annotationen, alleGene, atc3, atc4] */\n")
    f.write("const D_ROWS=%s;\n" % j(rows))
    f.write("const D_EFF=%s;\n" % j(EFFS))
    f.write("const D_EFFDE=%s;\n" % j([EFF_DE[e] for e in EFFS[:4]]+["Enzymatische Beeinflussung"]))
    f.write("const D_RISK=%s;\n" % j(subjects))
    f.write("/* [a, b, Schweregrad, Enzym, Wirkungsart, Risikobegriff, 0=a loest aus 1=b] */\n")
    f.write("const D_IX=%s;\n" % j(ix))
    f.write("/* [Wirkstoff, ...Alternativen derselben ATC-Ebene-4-Gruppe] */\n")
    f.write("const D_ALT=%s;\n" % j(alt))
    f.write("const D_RECGENES=%s;\n" % j(recgenes))
    f.write("/* [Wirkstoff, Gen, Stufe(-1=Sondergenotyp), RohGenotyp, Textindex, Leitlinienmaske] */\n")
    f.write("const D_REC=%s;\n" % j(rectab))

import os
print("\ngeschrieben:", OUT, "%.0f kB" % (os.path.getsize(OUT)/1024))
print("Wirkstoffe %d | mit Gen %d | mit Empfehlung %d | mit Leitlinie %d | Wechselwirkungen %d"
      % (len(rows),
         sum(1 for r in rows if r[3] >= 0),
         sum(1 for r in rows if r[7] >= 0),
         sum(1 for r in rows if r[6]),
         len(ix)))
