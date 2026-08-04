# -*- coding: utf-8 -*-
"""
Zwei-Gen-Zeilen im Spreadsheet. Eine Zeile mit GENE(1) und GENE (2) gilt nur,
wenn BEIDE Phaenotypen passen. Mein erster Abgleich hat nur GENE(1) geprueft -
dadurch wirkten drei Amitriptylin-Zeilen als Treffer, obwohl sie CYP2D6
voraussetzen, das in diesem VCF nicht gerufen werden kann.
"""
import json, os
from collections import Counter, defaultdict
from openpyxl import load_workbook

B = r"C:\Users\DanielWallerstorfer\Novogenia GmbH\AI RESOURCES - Dokumente\PHARMACOGENETICS"
D = r"C:\Users\DanielWallerstorfer\Downloads"
def kopf(t): print("\n" + "=" * 74 + "\n" + t + "\n" + "=" * 74)

wb = load_workbook(B + r"\Pharmgkb drug recommendations V4.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[0]
it = ws.iter_rows(values_only=True)
h = list(next(it))
print("Alle Spalten mit Index:")
for i, n in enumerate(h): print("  %2d %s" % (i, n))

rows = []
for row in it:
    if not row[0]: continue
    rows.append(row)
wb.close()

kopf("Alle Amitriptylin-Zeilen, roh")
for r in rows:
    if str(r[0]).strip().lower().startswith("amitript"):
        print("-" * 74)
        for i, n in enumerate(h):
            if r[i] not in (None, ""): print("  %-22s %s" % (str(n)[:22], str(r[i])[:150]))

kopf("Wie viele Zeilen haben zwei Gene?")
i_g1, i_gt1 = h.index("GENE(1)"), h.index("GENOTYPE (1)")
i_g2 = h.index("GENE (2)"); i_gt2 = h.index("METABOLIZER (2)")
ein = zwei = 0
paare = Counter()
for r in rows:
    if r[i_g2] not in (None, ""):
        zwei += 1; paare[(str(r[i_g1]).strip(), str(r[i_g2]).strip())] += 1
    else: ein += 1
print("  nur GENE(1):        %3d" % ein)
print("  GENE(1)+GENE (2):   %3d" % zwei)
for k, v in paare.most_common(): print("      %s + %s : %d" % (k[0], k[1], v))

kopf("Welche Wirkstoffe sind reine Ein-Gen-Zeilen?")
d1 = defaultdict(list)
for r in rows:
    if r[i_g2] in (None, ""):
        d1[str(r[i_g1]).strip()].append((str(r[0]).strip(), str(r[i_gt1]).strip()))
for g in sorted(d1):
    print("  %-9s %s" % (g, ", ".join("%s/%s" % x for x in sorted(set(d1[g])))))
