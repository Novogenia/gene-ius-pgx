# -*- coding: utf-8 -*-
"""
Kann die PharmCAT-Ausgabe gegen die Novogenia-Spreadsheets gematcht werden?

Schluessel waere (Gen, Phaenotyp). PharmCAT liefert Gen + Phaenotyp-String,
das Spreadsheet 'Pharmgkb drug recommendations V4' liefert Gen + GENOTYPE(1).
Diese Datei prueft ohne etwas zu bauen, wie weit die beiden Vokabulare passen.
"""
import json, os, re
from collections import defaultdict, Counter
from openpyxl import load_workbook

B = r"C:\Users\DanielWallerstorfer\Novogenia GmbH\AI RESOURCES - Dokumente\PHARMACOGENETICS"
D = r"C:\Users\DanielWallerstorfer\Downloads"

def kopf(t): print("\n" + "=" * 74 + "\n" + t + "\n" + "=" * 74)

# ---------------------------------------------- 1. PharmCAT: Gen -> Phaenotyp
pc = json.load(open(os.path.join(D, "pharmcat.phenotype (1).json"), encoding="utf-8"))
gr = pc["geneReports"]
pc_gene = {}
for g, r in gr.items():
    dl = r.get("recommendationDiplotypes") or []
    phen = sorted({p for d in dl for p in (d.get("phenotypes") or [])})
    labs = [d.get("label") for d in dl if d.get("label")]
    asc = sorted({str(d.get("activityScore")) for d in dl
                  if d.get("activityScore") not in (None, "", "n/a")})
    lookup = sorted({k for d in dl for k in (d.get("lookupKey") or [])})
    pc_gene[g] = dict(phen=phen, labels=labs, mehrdeutig=len(dl) > 1,
                      score=asc, lookup=lookup,
                      fehlend=len(r.get("uncalledHaplotypes") or []))

kopf("1. PharmCAT-Vokabular: Gen -> Phaenotyp (lookupKey)")
for g in sorted(pc_gene):
    v = pc_gene[g]
    print("%-9s %-40s %s%s" % (
        g, " | ".join(v["lookup"]) or "-",
        (v["labels"][0] if v["labels"] else "-"),
        "  [%d Diplotypen mehrdeutig]" % len(v["labels"]) if v["mehrdeutig"] else ""))

# ---------------------------------------------- 2. Spreadsheet-Vokabular
kopf("2. Spreadsheet 'Pharmgkb drug recommendations V4': GENOTYPE(1) je Gen")
wb = load_workbook(B + r"\Pharmgkb drug recommendations V4.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[0]
it = ws.iter_rows(values_only=True)
h = list(next(it))
print("Spalten:", [x for x in h if x])
ci = {n: h.index(n) for n in h if isinstance(n, str)}

sheet = defaultdict(Counter)      # Gen -> Counter(GENOTYPE)
sheet2 = defaultdict(Counter)     # Gen(2) -> Counter(METABOLIZER(2))
zeilen = []
for row in it:
    dn = row[ci["DRUG"]] if "DRUG" in ci else None
    if not dn: continue
    g1 = str(row[ci["GENE(1)"]]).strip() if row[ci.get("GENE(1)", -1)] else None
    gt1 = str(row[ci["GENOTYPE (1)"]]).strip() if row[ci.get("GENOTYPE (1)", -1)] else ""
    g2 = str(row[ci["GENE (2)"]]).strip() if ci.get("GENE (2)") is not None and row[ci["GENE (2)"]] else None
    gt2 = str(row[ci["METABOLIZER (2)"]]).strip() if ci.get("METABOLIZER (2)") is not None and row[ci["METABOLIZER (2)"]] else ""
    if g1: sheet[g1][gt1] += 1
    if g2: sheet2[g2][gt2] += 1
    zeilen.append((str(dn).strip(), g1, gt1, g2, gt2))
wb.close()
print("\nEmpfehlungszeilen:", len(zeilen))
for g in sorted(sheet):
    print("  %-10s %s" % (g, ", ".join("%s(%d)" % (k or "LEER", v)
                                        for k, v in sheet[g].most_common())))
print("\nGEN(2) / METABOLIZER(2):")
for g in sorted(sheet2):
    print("  %-10s %s" % (g, ", ".join("%s(%d)" % (k or "LEER", v)
                                        for k, v in sheet2[g].most_common())))

# ---------------------------------------------- 3. Uebersetzungstabelle pruefen
kopf("3. Passt PharmCAT-Phaenotyp auf Spreadsheet-GENOTYPE?")
# Mapping-Vorschlag: beides auf einen kanonischen Code bringen
CANON = {
  "ultrarapid metabolizer": "UM", "rapid metabolizer": "RM",
  "normal metabolizer": "NM", "extensive": "NM", "normal": "NM",
  "intermediate metabolizer": "IM", "intermediate": "IM",
  "poor metabolizer": "PM", "poor": "PM",
  "ultrarapid": "UM", "rapid": "RM",
  "normal function": "NF", "decreased function": "DF",
  "poor function": "PF", "increased function": "IF",
  "indeterminate": "?", "no result": "?", "n/a": "?",
  "uncertain susceptibility": "US", "malignant hyperthermia susceptibility": "MHS",
}
def canon(s):
    s = (s or "").strip().lower()
    return CANON.get(s, s.upper() if s else "")

sheet_codes = {g: {canon(k) for k in sheet[g] if k} for g in sheet}
print("%-10s %-22s %-22s %s" % ("GEN", "PharmCAT", "im Spreadsheet?", "Bemerkung"))
print("-" * 74)
treffer = fehl = 0
for g in sorted(pc_gene):
    v = pc_gene[g]
    if not v["lookup"]:
        print("%-10s %-22s %-22s %s" % (g, "(kein Ergebnis)", "-", "PharmCAT hat nichts gerufen"))
        continue
    for ph in v["lookup"]:
        c = canon(ph)
        if g in sheet_codes:
            ok = c in sheet_codes[g]
            print("%-10s %-22s %-22s %s" % (g, ph, "JA" if ok else "NEIN",
                  "" if ok else "Spreadsheet hat: " + ", ".join(sorted(sheet_codes[g]))))
            treffer += ok; fehl += (not ok)
        else:
            print("%-10s %-22s %-22s %s" % (g, ph, "Gen fehlt", "kein Wirkstoff-Bezug im Spreadsheet"))
            fehl += 1
print("\nphaenotypische Treffer: %d | Fehlschlaege: %d" % (treffer, fehl))

# ---------------------------------------------- 4. Wieviele Wirkstoffe treffen?
kopf("4. Konkrete Wirkstoff-Bewertungen aus diesem Genprofil")
prof = {}
for g, v in pc_gene.items():
    if v["lookup"] and not v["mehrdeutig"]:
        prof[g] = canon(v["lookup"][0])
    elif v["lookup"] and v["mehrdeutig"]:
        cs = {canon(x) for x in v["lookup"]}
        prof[g] = ("MEHRDEUTIG", sorted(cs))
print("eindeutiges Profil:", {k: v for k, v in prof.items() if not isinstance(v, tuple)})
print("mehrdeutig:", {k: v[1] for k, v in prof.items() if isinstance(v, tuple)})

eind = {k: v for k, v in prof.items() if not isinstance(v, tuple)}
getroffen = defaultdict(list)
for dn, g1, gt1, g2, gt2 in zeilen:
    if not g1: continue
    if g1 in eind and canon(gt1) == eind[g1]:
        getroffen[dn].append((g1, gt1))
print("\nWirkstoffe mit einer passenden Empfehlungszeile:", len(getroffen))
for dn in sorted(getroffen)[:40]:
    print("   %-34s %s" % (dn, "; ".join("%s %s" % x for x in getroffen[dn])))
if len(getroffen) > 40: print("   ... und %d weitere" % (len(getroffen) - 40))
